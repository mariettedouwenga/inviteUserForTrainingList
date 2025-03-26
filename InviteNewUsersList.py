import pandas as pd
import streamlit as st
import openpyxl as px
import markdown as md

header = st.container()
dataset = st.container()

with header:
    st.title("Infor LN New Users")
    st.image("userInvite.png")
    st.divider()

file_path = "NewUsers_2025.xlsx"
training_names = pd.read_excel("trainingName.xlsx")

#loop through this sheet to find column name selected in the selection box
training_name_sh = file_path.active

#Select the training name
column_name = st.selectbox('Select the training name', training_names, index = None,
    placeholder="Select training name",)

#Input to enter the date on which the training will take place
training_date = st.date_input("Enter the training date")

def find_first_empty_row(file_path, column_name):
    # Read the Excel file
    df = pd.read_excel(file_path)
    
    # Check if the column exists
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' does not exist in the dataframe.")
    
    # Get the column data
    column_data = df[column_name]
    
    # Find the first empty row in the column
    first_empty_row = column_data[column_data.isna()].index[0]
    #st.write(str(training_date.value))
    
    return first_empty_row

#Validation of first empty row in a specific Column
try:
    first_empty_row = find_first_empty_row(file_path, column_name)
    print(f"The first empty row in column '{column_name}' is: {first_empty_row}")
except ValueError as e:
    print(e)
except IndexError:
    print(f"No empty rows found in column '{column_name}'")




#Add the training_date to the first empty row of the training name column
for j in range(1, training_name_sh.max_column+1): 

        #Iterate through rows in excel sheet
        for i in range(1, training_name_sh.max_row+1): 
            cell_obj = training_name_sh.cell(1, column=i)
                 
            if cell_obj.value == column_name.value:
                if training_name_sh.cell(i,j).value == "":
                    training_name_sh.cell(i,j).value == training_date

                                                                          
                                                
                   
    